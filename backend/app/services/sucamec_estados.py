from __future__ import annotations

import json
import os
import re
import signal
import subprocess
import sys
import time
import threading
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from typing import Literal

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[2]
DEFAULT_FLOW_DIR = BASE_DIR / "ESTADOS-GADSO"
ALLOWED_GROUPS = {"JV", "SELVA", "TODOS"}
PREVIEW_ROWS = 5
UPLOAD_PREFIX = "entrada_"
KEEP_INPUT_FILENAMES = {"plantilla_mis_vigilantes.xlsx"}
REQUIRED_INPUT_COLUMNS = {"NRO DOCUMENTO"}
OPTIONAL_INPUT_COLUMNS = {"NOMBRE"}
ALLOWED_INPUT_COLUMNS = REQUIRED_INPUT_COLUMNS | OPTIONAL_INPUT_COLUMNS

JobStatus = Literal["queued", "running", "success", "error", "cancelled"]


@dataclass
class SucamecJob:
    id: str
    grupo: str
    input_filename: str
    status: JobStatus = "queued"
    message: str = "Pendiente"
    created_at: datetime = field(default_factory=datetime.now)
    started_at: datetime | None = None
    finished_at: datetime | None = None
    return_code: int | None = None
    pid: int | None = None
    input_path: str | None = None
    expected_run_name: str | None = None
    expected_output_dir: str | None = None
    output_dir: str | None = None
    result_file: str | None = None
    result_files: list[str] = field(default_factory=list)
    log_tail: str = ""


_jobs: dict[str, SucamecJob] = {}
_processes: dict[str, subprocess.Popen] = {}
_lock = threading.Lock()


def _resolve_backend_path(value: str | None, default: Path) -> Path:
    path = Path(value).expanduser() if value else default
    if not path.is_absolute():
        path = BASE_DIR / path
    return path.resolve()


def _flow_dir() -> Path:
    return _resolve_backend_path(os.getenv("ESTADOS_GADSO_DIR"), DEFAULT_FLOW_DIR)


def _input_dir() -> Path:
    return _flow_dir() / "data" / "entrada_data"


def _lots_dir() -> Path:
    return _flow_dir() / "lotes"


def _runtime_dir() -> Path:
    return _flow_dir() / "runtime"


def _jobs_dir() -> Path:
    return _runtime_dir() / "jobs"


def _locks_dir() -> Path:
    return _runtime_dir() / "locks"


def _estados_lock_path() -> Path:
    return _locks_dir() / "estados-carne.lock"


def _python_executable() -> str:
    configured_python = os.getenv("ESTADOS_GADSO_PYTHON", "").strip()
    if configured_python:
        return configured_python

    for relative_python in (("Scripts", "python.exe"), ("bin", "python")):
        flow_venv_python = _flow_dir() / ".venv" / Path(*relative_python)
        if flow_venv_python.exists():
            return str(flow_venv_python)

    return sys.executable


def _max_upload_bytes() -> int:
    raw_value = os.getenv("ESTADOS_GADSO_MAX_UPLOAD_MB", "50").strip()
    try:
        megabytes = max(1, int(raw_value))
    except ValueError:
        megabytes = 50
    return megabytes * 1024 * 1024


def _max_retained_uploads() -> int:
    raw_value = os.getenv("ESTADOS_GADSO_MAX_RETAINED_UPLOADS", "0").strip()
    try:
        return max(0, int(raw_value))
    except ValueError:
        return 0


def _upload_ttl_seconds() -> int:
    raw_value = os.getenv("ESTADOS_GADSO_UPLOAD_TTL_MINUTES", "60").strip()
    try:
        minutes = max(1, int(raw_value))
    except ValueError:
        minutes = 60
    return minutes * 60


def _max_retained_jobs() -> int:
    raw_value = os.getenv("ESTADOS_GADSO_MAX_RETAINED_JOBS", "10").strip()
    try:
        return max(1, int(raw_value))
    except ValueError:
        return 10


def _max_concurrent_jobs() -> int:
    raw_value = os.getenv("ESTADOS_GADSO_MAX_CONCURRENT_JOBS", "1").strip()
    try:
        return max(1, int(raw_value))
    except ValueError:
        return 1


def _job_timeout_seconds() -> int:
    raw_value = os.getenv("ESTADOS_GADSO_JOB_TIMEOUT_MINUTES", "120").strip()
    try:
        minutes = max(1, int(raw_value))
    except ValueError:
        minutes = 120
    return minutes * 60


def ensure_flow_paths() -> None:
    if not (_flow_dir() / "src" / "agents_flow").exists():
        raise FileNotFoundError(f"No se encontro el flujo ESTADOS-GADSO en {_flow_dir()}")

    _input_dir().mkdir(parents=True, exist_ok=True)
    _lots_dir().mkdir(parents=True, exist_ok=True)
    _jobs_dir().mkdir(parents=True, exist_ok=True)
    _locks_dir().mkdir(parents=True, exist_ok=True)
    _cleanup_expired_uploads()


def _serialize_datetime(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def _job_to_payload(job: SucamecJob) -> dict:
    payload = asdict(job)
    payload["created_at"] = _serialize_datetime(job.created_at)
    payload["started_at"] = _serialize_datetime(job.started_at)
    payload["finished_at"] = _serialize_datetime(job.finished_at)
    return payload


def _job_from_payload(payload: dict) -> SucamecJob:
    return SucamecJob(
        id=str(payload["id"]),
        grupo=str(payload["grupo"]),
        input_filename=str(payload["input_filename"]),
        status=payload.get("status", "queued"),
        message=payload.get("message", "Pendiente"),
        created_at=datetime.fromisoformat(payload["created_at"]),
        started_at=datetime.fromisoformat(payload["started_at"]) if payload.get("started_at") else None,
        finished_at=datetime.fromisoformat(payload["finished_at"]) if payload.get("finished_at") else None,
        return_code=payload.get("return_code"),
        pid=payload.get("pid"),
        input_path=payload.get("input_path"),
        expected_run_name=payload.get("expected_run_name"),
        expected_output_dir=payload.get("expected_output_dir"),
        output_dir=payload.get("output_dir"),
        result_file=payload.get("result_file"),
        result_files=list(payload.get("result_files") or ([] if not payload.get("result_file") else [payload["result_file"]])),
        log_tail=payload.get("log_tail", ""),
    )


def _job_path(job_id: str) -> Path:
    return _jobs_dir() / f"{job_id}.json"


def _save_job(job: SucamecJob) -> None:
    ensure_flow_paths()
    _job_path(job.id).write_text(json.dumps(_job_to_payload(job), ensure_ascii=False, indent=2), encoding="utf-8")
    _cleanup_old_jobs(keep_job_id=job.id)


def _cleanup_old_jobs(keep_job_id: str | None = None) -> int:
    jobs_dir = _jobs_dir()
    if not jobs_dir.exists():
        return 0

    try:
        files = [path for path in jobs_dir.glob("*.json") if path.is_file()]
    except Exception:
        return 0

    keep_total = _max_retained_jobs()
    if len(files) <= keep_total:
        return 0

    protected_name = f"{keep_job_id}.json" if keep_job_id else ""
    files.sort(key=lambda path: path.stat().st_mtime, reverse=True)

    deleted = 0
    for path in reversed(files):
        if path.name == protected_name:
            continue
        try:
            path.unlink(missing_ok=True)
            deleted += 1
        except Exception:
            continue
        if len(files) - deleted <= keep_total:
            break

    return deleted


def _load_job(job_id: str) -> SucamecJob | None:
    path = _job_path(job_id)
    if not path.exists():
        return None

    try:
        return _job_from_payload(json.loads(path.read_text(encoding="utf-8")))
    except (KeyError, ValueError, json.JSONDecodeError):
        return None


def _store_job(job: SucamecJob) -> None:
    with _lock:
        _jobs[job.id] = job
        _save_job(job)


def _store_process(job_id: str, process: subprocess.Popen) -> None:
    with _lock:
        _processes[job_id] = process


def _pop_process(job_id: str) -> subprocess.Popen | None:
    with _lock:
        return _processes.pop(job_id, None)


def _get_process(job_id: str) -> subprocess.Popen | None:
    with _lock:
        return _processes.get(job_id)


def _public_job(job: SucamecJob) -> dict:
    return {
        "id": job.id,
        "grupo": job.grupo,
        "input_filename": job.input_filename,
        "display_input_filename": _display_input_filename(job.input_filename),
        "status": job.status,
        "message": job.message,
        "created_at": job.created_at.isoformat(),
        "started_at": job.started_at.isoformat() if job.started_at else None,
        "finished_at": job.finished_at.isoformat() if job.finished_at else None,
        "return_code": job.return_code,
        "pid": job.pid,
        "expected_run_name": job.expected_run_name,
        "has_result": bool(job.result_file),
        "result_files": [Path(path).name for path in job.result_files],
        "log_tail": job.log_tail[-4000:],
    }


def _display_input_filename(filename: str) -> str:
    safe_name = Path(filename or "").name
    return re.sub(r"^entrada_\d{8}_\d{6}_", "", safe_name)


def _resolve_input_file(filename: str) -> Path:
    ensure_flow_paths()
    safe_name = Path(filename or "").name
    if not safe_name.lower().endswith(".xlsx"):
        raise ValueError("Archivo de entrada invalido")

    input_dir = _input_dir().resolve()
    path = (input_dir / safe_name).resolve()
    try:
        path.relative_to(input_dir)
    except ValueError as exc:
        raise ValueError("Archivo de entrada invalido") from exc

    if not path.exists() or not path.is_file():
        raise ValueError("El archivo de entrada no existe o ya no esta disponible")

    return path


def _is_temporary_upload(path: Path) -> bool:
    return path.name.startswith(UPLOAD_PREFIX) and path.suffix.lower() == ".xlsx"


def _cleanup_old_uploads(keep_filename: str | None = None) -> None:
    _cleanup_expired_uploads(keep_filename=keep_filename)
    keep_name = Path(keep_filename).name if keep_filename else ""
    max_retained = _max_retained_uploads()
    candidates = [
        path
        for path in _input_dir().glob(f"{UPLOAD_PREFIX}*.xlsx")
        if path.is_file() and path.name != keep_name and path.name not in KEEP_INPUT_FILENAMES
    ]
    candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)

    for path in candidates[max_retained:]:
        path.unlink(missing_ok=True)


def _cleanup_expired_uploads(keep_filename: str | None = None) -> None:
    input_dir = _input_dir()
    if not input_dir.exists():
        return

    keep_name = Path(keep_filename).name if keep_filename else ""
    expires_before = time.time() - _upload_ttl_seconds()

    for path in input_dir.glob(f"{UPLOAD_PREFIX}*.xlsx"):
        if not path.is_file() or path.name == keep_name or path.name in KEEP_INPUT_FILENAMES:
            continue
        try:
            if path.stat().st_mtime < expires_before:
                path.unlink(missing_ok=True)
        except OSError:
            continue


def _cleanup_job_input(input_path: Path) -> None:
    if _is_temporary_upload(input_path) and input_path.name not in KEEP_INPUT_FILENAMES:
        input_path.unlink(missing_ok=True)


def _active_jobs_count() -> int:
    active_statuses = {"queued", "running"}
    active_count = 0

    with _lock:
        active_count += sum(1 for job in _jobs.values() if job.status in active_statuses)

    jobs_dir = _jobs_dir()
    if not jobs_dir.exists():
        return active_count

    seen_job_ids = set()
    with _lock:
        seen_job_ids.update(_jobs.keys())

    for path in jobs_dir.glob("*.json"):
        job_id = path.stem
        if job_id in seen_job_ids:
            continue
        job = _load_job(job_id)
        if job and job.status in active_statuses:
            active_count += 1

    return active_count


def _ensure_concurrency_capacity() -> None:
    ensure_flow_paths()
    max_concurrent = _max_concurrent_jobs()
    if _active_jobs_count() >= max_concurrent:
        raise ValueError(
            "Ya existe una ejecucion SUCAMEC en curso. "
            "Espera a que finalice o ajusta ESTADOS_GADSO_MAX_CONCURRENT_JOBS."
        )


def _write_lock_file(job: SucamecJob) -> None:
    lock_path = _estados_lock_path()
    lock_payload = {
        "job_id": job.id,
        "grupo": job.grupo,
        "pid": job.pid,
        "started_at": _serialize_datetime(job.started_at),
        "input_filename": job.input_filename,
    }
    lock_path.write_text(json.dumps(lock_payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _clear_lock_file(job_id: str) -> None:
    lock_path = _estados_lock_path()
    if not lock_path.exists():
        return

    try:
        payload = json.loads(lock_path.read_text(encoding="utf-8"))
        if str(payload.get("job_id", "")) != job_id:
            return
    except Exception:
        pass

    lock_path.unlink(missing_ok=True)


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().upper()
    return re.sub(r"\s+", " ", text)


def _cell_to_text_preserving_zeros(cell) -> str:
    value = cell.value
    if value is None:
        return ""

    if isinstance(value, str):
        return value.strip()

    number_format = str(cell.number_format or "")
    if isinstance(value, int):
        digits = str(value)
    elif isinstance(value, float) and value.is_integer():
        digits = str(int(value))
    else:
        return str(value).strip()

    if re.fullmatch(r"0+", number_format):
        return digits.zfill(len(number_format))
    return digits


def _validate_input_excel(path: Path) -> dict:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = workbook.active
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if not header_cells:
            raise ValueError("El Excel de entrada no tiene cabecera")

        headers = [_normalize_header(cell.value) for cell in header_cells]
        header_map = {header: index for index, header in enumerate(headers) if header}
        detected_columns = [header for header in headers if header]
        unexpected_columns = [header for header in detected_columns if header not in ALLOWED_INPUT_COLUMNS]
        missing_columns = [header for header in REQUIRED_INPUT_COLUMNS if header not in header_map]

        if missing_columns:
            raise ValueError(
                "Formato invalido. El Excel debe tener la columna obligatoria NRO DOCUMENTO y, opcionalmente, NOMBRE."
            )

        if unexpected_columns:
            raise ValueError(
                "Formato invalido. Solo se aceptan las columnas NRO DOCUMENTO y NOMBRE. "
                f"Columnas no permitidas: {', '.join(unexpected_columns)}."
            )

        document_index = header_map["NRO DOCUMENTO"]
        preview: list[dict[str, str | int]] = []
        omitted_rows: list[int] = []
        total_records = 0

        for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            document = _cell_to_text_preserving_zeros(row[document_index])
            if not document and not any(_cell_to_text_preserving_zeros(cell) for cell in row):
                continue
            if not document:
                omitted_rows.append(row_number)
                continue

            total_records += 1
            if len(preview) < PREVIEW_ROWS:
                preview.append(
                    {
                        "row_number": len(preview) + 1,
                        "nro_documento": document,
                    }
                )

        if total_records == 0:
            raise ValueError("El Excel no contiene registros con NRO DOCUMENTO para procesar")

        return {
            "validation": {
                "is_valid": True,
                "required_columns": sorted(REQUIRED_INPUT_COLUMNS),
                "detected_columns": detected_columns,
                "document_column": "NRO DOCUMENTO",
                "total_records": total_records,
                "omitted_rows": omitted_rows[:25],
                "omitted_rows_count": len(omitted_rows),
            },
            "preview": preview,
        }
    finally:
        workbook.close()


def save_input_excel(upload_file, original_filename: str) -> dict:
    ensure_flow_paths()
    filename = Path(original_filename or "entrada.xlsx").name
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("Solo se aceptan archivos .xlsx")

    target_name = f"entrada_{datetime.now():%Y%m%d_%H%M%S}_{filename}"
    target_path = _input_dir() / target_name
    max_bytes = _max_upload_bytes()
    total_bytes = 0

    try:
        with target_path.open("wb") as buffer:
            while chunk := upload_file.read(1024 * 1024):
                total_bytes += len(chunk)
                if total_bytes > max_bytes:
                    raise ValueError(f"El archivo supera el limite de {max_bytes // 1024 // 1024} MB")
                buffer.write(chunk)
    except Exception:
        target_path.unlink(missing_ok=True)
        raise

    try:
        inspection = _validate_input_excel(target_path)
    except Exception:
        target_path.unlink(missing_ok=True)
        raise

    _cleanup_old_uploads(keep_filename=target_name)

    return {"filename": target_name, "size_bytes": total_bytes, **inspection}


def create_job(grupo: str, input_filename: str) -> dict:
    grupo_normalized = grupo.strip().upper() or "JV"
    if grupo_normalized not in ALLOWED_GROUPS:
        raise ValueError("Grupo invalido")

    _ensure_concurrency_capacity()
    input_path = _resolve_input_file(input_filename)
    _validate_input_excel(input_path)
    job_id = uuid.uuid4().hex
    expected_run_name = f"api_{datetime.now():%Y%m%d_%H%M%S}_{job_id[:8]}"
    job = SucamecJob(
        id=job_id,
        grupo=grupo_normalized,
        input_filename=Path(input_filename).name,
        input_path=str(input_path),
        expected_run_name=expected_run_name,
        expected_output_dir=str(_lots_dir() / expected_run_name),
    )
    _store_job(job)

    thread = threading.Thread(target=_run_job, args=(job.id,), daemon=True)
    thread.start()
    return _public_job(job)


def get_job(job_id: str) -> dict | None:
    with _lock:
        job = _jobs.get(job_id)

    if not job:
        job = _load_job(job_id)

    return _public_job(job) if job else None


def get_result_file(job_id: str) -> Path | None:
    with _lock:
        job = _jobs.get(job_id)

    if not job:
        job = _load_job(job_id)

    if not job or not job.result_file:
        return None

    path = Path(job.result_file)
    if not path.exists() or not path.is_file():
        return None

    try:
        path.resolve().relative_to(_lots_dir().resolve())
    except ValueError:
        return None

    return path


def get_result_files(job_id: str) -> list[Path]:
    with _lock:
        job = _jobs.get(job_id)

    if not job:
        job = _load_job(job_id)

    if not job:
        return []

    raw_paths = job.result_files or []
    if job.output_dir:
        output_dir = Path(job.output_dir)
        if output_dir.exists() and output_dir.is_dir():
            raw_paths = [
                str(path)
                for path in sorted(
                    [
                        *output_dir.glob("RB_GADSOCarnetSUCAMEC_*.xlsx"),
                        *output_dir.glob("RB_GADSOValidacionNoEncontradosSUCAMEC_*.xlsx"),
                    ],
                    key=lambda item: item.stat().st_mtime,
                )
            ] or raw_paths

    if not raw_paths and job.result_file:
        raw_paths = [job.result_file]

    result_paths: list[Path] = []
    lots_dir = _lots_dir().resolve()

    for raw_path in raw_paths:
        if not raw_path:
            continue
        path = Path(raw_path)
        if not path.exists() or not path.is_file():
            continue
        try:
            path.resolve().relative_to(lots_dir)
        except ValueError:
            continue
        result_paths.append(path)

    return result_paths


def _latest_lot_dirs() -> set[Path]:
    lots_dir = _lots_dir()
    if not lots_dir.exists():
        return set()
    return {path.resolve() for path in lots_dir.iterdir() if path.is_dir()}


def _collect_result_files(output_dir: Path) -> tuple[str | None, str | None, list[str]]:
    if not output_dir.exists() or not output_dir.is_dir():
        return None, None, []

    result_files = sorted(
        [
            *output_dir.glob("RB_GADSOCarnetSUCAMEC_*.xlsx"),
            *output_dir.glob("RB_GADSOValidacionNoEncontradosSUCAMEC_*.xlsx"),
        ],
        key=lambda path: path.stat().st_mtime,
    )
    primary_files = [path for path in result_files if path.name.startswith("RB_GADSOCarnetSUCAMEC_")]
    result_file = primary_files[-1] if primary_files else (result_files[-1] if result_files else None)
    return str(output_dir), str(result_file) if result_file else None, [str(path) for path in result_files]


def _find_newest_result(before_dirs: set[Path], expected_output_dir: str | None = None) -> tuple[str | None, str | None, list[str]]:
    if expected_output_dir:
        expected_result = _collect_result_files(Path(expected_output_dir))
        if expected_result[1] or expected_result[2]:
            return expected_result

    lots_dir = _lots_dir()
    candidates = [path for path in lots_dir.iterdir() if path.is_dir() and path.resolve() not in before_dirs]
    if not candidates:
        candidates = [path for path in lots_dir.iterdir() if path.is_dir()]
    if not candidates:
        return None, None, []

    output_dir = max(candidates, key=lambda path: path.stat().st_mtime)
    return _collect_result_files(output_dir)


def _update_job(job_id: str, **changes) -> SucamecJob:
    with _lock:
        job = _jobs.get(job_id) or _load_job(job_id)
        if not job:
            raise ValueError("No se encontro la ejecucion solicitada")
        for key, value in changes.items():
            setattr(job, key, value)
        _jobs[job_id] = job
        _save_job(job)
        return job


def _terminate_process_tree(process: subprocess.Popen) -> None:
    if process.poll() is not None:
        return

    if os.name == "nt":
        subprocess.run(
            ["taskkill", "/PID", str(process.pid), "/T", "/F"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
        return

    try:
        os.killpg(os.getpgid(process.pid), signal.SIGTERM)
    except ProcessLookupError:
        return
    except Exception:
        process.terminate()


def cancel_job(job_id: str) -> dict:
    with _lock:
        job = _jobs.get(job_id)

    if not job:
        job = _load_job(job_id)

    if not job:
        raise ValueError("No se encontro la ejecucion solicitada")

    if job.status not in {"queued", "running"}:
        return _public_job(job)

    process = _get_process(job_id)
    if process:
        _terminate_process_tree(process)

    updated = _update_job(
        job_id,
        status="cancelled",
        message="Ejecucion cancelada por el usuario",
        finished_at=datetime.now(),
    )
    _clear_lock_file(job_id)
    return _public_job(updated)


def _run_job(job_id: str) -> None:
    current_job = _load_job(job_id)
    if current_job and current_job.status == "cancelled":
        return

    job = _update_job(
        job_id,
        status="running",
        message="Ejecutando flujo SUCAMEC",
        started_at=datetime.now(),
    )

    input_path = _resolve_input_file(job.input_filename)
    before_dirs = _latest_lot_dirs()
    command = [_python_executable(), "-m", "src.agents_flow.cli", "--grupo", job.grupo]
    env = os.environ.copy()
    env.setdefault("CARNET_HEADLESS", "1")
    env.setdefault("SUCAMEC_HEADLESS", "1")
    env.setdefault("HOLD_BROWSER_OPEN", "0")
    env["SUCAMEC_INPUT_EXCEL"] = str(input_path)
    if job.expected_run_name:
        env["SUCAMEC_RUN_NAME"] = job.expected_run_name

    output_tail: list[str] = []

    try:
        popen_kwargs: dict[str, Any] = {}
        if os.name == "nt":
            popen_kwargs["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP
        else:
            popen_kwargs["start_new_session"] = True

        process = subprocess.Popen(
            command,
            cwd=str(_flow_dir()),
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            **popen_kwargs,
        )
        _store_process(job_id, process)
        _update_job(job_id, pid=process.pid, input_path=str(input_path), expected_output_dir=job.expected_output_dir)
        _write_lock_file(_load_job(job_id) or job)

        assert process.stdout is not None
        timed_out = False

        def _mark_timeout_and_terminate() -> None:
            nonlocal timed_out
            timed_out = True
            _terminate_process_tree(process)

        timeout_timer = threading.Timer(_job_timeout_seconds(), _mark_timeout_and_terminate)
        timeout_timer.daemon = True
        timeout_timer.start()

        for line in process.stdout:
            output_tail.append(line.rstrip())
            if len(output_tail) > 200:
                output_tail = output_tail[-200:]
            _update_job(job_id, log_tail="\n".join(output_tail))

        return_code = process.wait()
        timeout_timer.cancel()
        _pop_process(job_id)
        current_job = _load_job(job_id)
        if current_job and current_job.status == "cancelled":
            _cleanup_job_input(input_path)
            _cleanup_old_uploads()
            _clear_lock_file(job_id)
            return

        output_dir, result_file, result_files = _find_newest_result(before_dirs, job.expected_output_dir)
        status: JobStatus = "success" if return_code == 0 else "error"
        message = "Flujo finalizado correctamente"
        if timed_out:
            output_tail.append("Timeout operativo alcanzado; proceso cancelado por el backend.")
            message = "El flujo excedio el tiempo maximo configurado"
        if return_code == 0 and not result_file:
            message = "Flujo finalizado, pero no se encontro Excel de salida"
        elif return_code != 0:
            message = message if "tiempo maximo" in message else "El flujo termino con error"

        _update_job(
            job_id,
            status=status,
            message=message,
            return_code=return_code,
            finished_at=datetime.now(),
            output_dir=output_dir,
            result_file=result_file,
            result_files=result_files,
            log_tail="\n".join(output_tail),
        )
        _cleanup_job_input(input_path)
        _cleanup_old_uploads()
        _clear_lock_file(job_id)
    except Exception as exc:
        _pop_process(job_id)
        current_job = _load_job(job_id)
        if current_job and current_job.status == "cancelled":
            _cleanup_job_input(input_path)
            _cleanup_old_uploads()
            return

        _update_job(
            job_id,
            status="error",
            message=f"No se pudo ejecutar el flujo: {exc}",
            finished_at=datetime.now(),
            log_tail="\n".join(output_tail),
        )
        _cleanup_job_input(input_path)
        _cleanup_old_uploads()
        _clear_lock_file(job_id)
